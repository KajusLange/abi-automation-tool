# excel_handler.py
from openpyxl import load_workbook

wb = load_workbook(filename="template_excel.xlsx")
ws = wb.active


def create_student_xlsx(student):
    # student info
    ws["B70"] = student.klasse
    ws["D70"] = student.name
    ws["G70"] = student.vorname
    ws["D73"] = student.pb
    ws["G73"] = student.wdh

    # final step
    wb.save(filename=f"{student.klasse}_{student.name}_{student.vorname}")


def adjust_row_count(groups):
    ws.row_dimensions.group(12 + groups[0], 21, hidden=True)
    ws.row_dimensions.group(25 + groups[1], 34, hidden=True)
    ws.row_dimensions.group(38 + groups[2], 47, hidden=True)
    ws.row_dimensions.group(51 + groups[3], 65, hidden=True)


def write_kursbelegung(kurse):
    start_index = (12, 25, 38, 51)
    group_size = (3, 3, 3, 1)

    for g in range(4):
        group = [x for x in kurse if x.agf == g + 1]
        for row in range(start_index[g], start_index[g] + group_size[g]):
            k = group[row-start_index[g]]

            if k.abi == 1 or k.abi == 2:
                ws[f"E{row}"] = "LK"
            elif k.diffkurs is True:
                ws[f"E{row}"] = "DK"
            else:
                ws[f"E{row}"] = "GK"

            ws[f"B{row}"] = k.title
            ws[f"D{row}"] = k.abi

            for column in range(6, 10):
                ws.cell(row=row, column=column).value = k.points[column-6]

            ws[f"S{row}"] = k.points[4]
            ws[f"U{row}"] = k.points[5]
